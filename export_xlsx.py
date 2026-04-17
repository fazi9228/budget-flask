"""
Finance-format XLSX export matching the APAC Marketing FY26 Tracker layout.
69 columns: A-BQ

GROUPING (Finance-facing):
  Primary grouping = channel_name (original behaviour, matches Finance template)
  Special case: if channel_name == "Performance Marketing" (the umbrella introduced
  in the refactor), rows are split by finance_cat ("Paid Social-Meta", "Bing",
  "Paid Social-Douyin", "Local Direct Deals", etc.) so Finance still sees the
  breakdown they expect. All other channels (Affiliate - CPA & FF,
  Campaign/Promotions, Events, Influencer/KOL, etc.) group by channel_name
  exactly as before.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from collections import OrderedDict

# The umbrella channel from the refactor. Rows in this channel are split
# by finance_cat so Finance still sees Meta/Bing/Douyin/LDD as separate lines.
PM_UMBRELLA = "Performance Marketing"

# ── COLOURS ──────────────────────────────────────────────────
C_DARK_GREEN  = "1F4E39"
C_MID_GREEN   = "375623"
C_LIGHT_GREEN = "E2EFDA"
C_DARK_GREY   = "404040"
C_WHITE       = "FFFFFF"
C_BLUE_INPUT  = "0070C0"
C_ORANGE      = "C65911"

MONTH_DATES = [
    datetime(2025,7,31), datetime(2025,8,31), datetime(2025,9,30),
    datetime(2025,10,31), datetime(2025,11,30), datetime(2025,12,31),
    datetime(2026,1,31), datetime(2026,2,28), datetime(2026,3,31),
    datetime(2026,4,30), datetime(2026,5,31), datetime(2026,6,30),
]
MONTH_KEYS = ["2025-07","2025-08","2025-09","2025-10","2025-11","2025-12",
              "2026-01","2026-02","2026-03","2026-04","2026-05","2026-06"]

COL_ACCOUNT=1; COL_OWNER=2; COL_CATEGORY=3; COL_COUNTRY=4; COL_VENDOR=5; COL_NOTE=6
COL_FY26_ACT=7; COL_FY26_BUD=8; COL_VAR_CHECK=9
COL_YTD_BUD=11; COL_YTD_ACT=12; COL_YTD_VAR=13
COL_CM_BUD=15; COL_CM_ACT=16; COL_CM_VAR=17
COL_BUD_START=19; COL_ACT_START=32; COL_VAR_START=45; COL_INP_START=58
TOTAL_COLS=69
NUM_FMT='#,##0'
gl=get_column_letter

def _fill(h): return PatternFill("solid",fgColor=h)
def _font(b=False,c="000000",s=10,i=False): return Font(name="Calibri",bold=b,color=c,size=s,italic=i)
def _border(c="CCCCCC"):
    s=Side(style='thin',color=c); return Border(left=s,right=s,top=s,bottom=s)
def _s(c,bg=None,bold=False,fc="000000",size=10,align="left",wrap=False,nf=None,italic=False):
    if bg: c.fill=_fill(bg)
    c.font=_font(bold,fc,size,italic)
    c.alignment=Alignment(horizontal=align,vertical='center',wrap_text=wrap)
    if nf: c.number_format=nf


def _group_key(entry):
    """Return the Finance-facing grouping label for an entry.
    - PM umbrella entries → split by finance_cat (falls back to 'Performance Marketing - Other')
    - Everything else → grouped by channel_name (matches original Finance template)
    """
    ch = str(entry.get("channel_name", "") or "").strip()
    if ch == PM_UMBRELLA:
        fc = str(entry.get("finance_cat", "") or "").strip()
        return fc if fc else f"{PM_UMBRELLA} - Other"
    return ch or entry.get("finance_cat") or "Other"


def build_finance_export(output_path, entries, channels, budgets, current_month_key="2026-02"):
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="APAC"
    ws.sheet_view.showGridLines=False

    for col,w in {1:30,2:16,3:28,4:14,5:22,6:28,7:16,8:14,9:12}.items():
        ws.column_dimensions[gl(col)].width=w
    for i in range(12):
        for base in [COL_BUD_START,COL_ACT_START,COL_VAR_START,COL_INP_START]:
            ws.column_dimensions[gl(base+i)].width=12
    for sp in [10,14,18,31,44,57]:
        ws.column_dimensions[gl(sp)].width=2

    for r,(txt,bg,fc,sz) in enumerate([
        ("Marketing Expenses Detail— APAC",C_DARK_GREEN,C_WHITE,12),
        ("Marketing FY26 Tracker",None,"000000",10),
        ("Pepperstone Group Limited",None,"000000",10)],1):
        ws.merge_cells(f'A{r}:{gl(TOTAL_COLS)}{r}')
        _s(ws.cell(r,1,txt),bg=bg,bold=True,fc=fc,size=sz)

    ws['C4']='AUD $000'; _s(ws['C4'],bg=C_DARK_GREEN,bold=True,fc=C_WHITE,align='center')
    ws['D4']='FY26'
    cm_idx=MONTH_KEYS.index(current_month_key) if current_month_key in MONTH_KEYS else 7
    ws['C5']='Current mth'; ws['D5']=MONTH_DATES[cm_idx]; ws['D5'].number_format='MMM-YY'

    for col,lbl in [(COL_OWNER,"APAC Total"),(COL_CATEGORY,"APAC")]:
        _s(ws.cell(7,col,lbl),bg=C_DARK_GREEN,bold=True,fc=C_WHITE)
    for col,lbl in [(COL_YTD_BUD,"Budget"),(COL_YTD_ACT,"Actual"),(COL_YTD_VAR,"Var"),
                     (COL_CM_BUD,"Budget"),(COL_CM_ACT,"Actual"),(COL_CM_VAR,"Var")]:
        _s(ws.cell(7,col,lbl),bg=C_DARK_GREEN,bold=True,fc=C_WHITE,size=9,align='center')
    for i in range(12):
        for base,lbl,bg in [(COL_BUD_START,"Budget",C_DARK_GREEN),(COL_ACT_START,"Actual",C_DARK_GREY),
                             (COL_VAR_START,"Var",C_DARK_GREY),(COL_INP_START,"Actual",C_ORANGE)]:
            _s(ws.cell(7,base+i,lbl),bg=bg,bold=True,fc=C_WHITE,size=8,align='center')

    for col,lbl in [(1,"AP- Main Account"),(3,"Category"),(4,"Country"),(5,"Vendor"),
                     (6,"Note"),(7,"FY26 (ACT/BUDGET)"),(8,"Budget"),(9,"Var Check"),
                     (11,"YTD"),(12,"YTD"),(13,"YTD")]:
        c=ws.cell(8,col,lbl); _s(c,bg=C_DARK_GREY,bold=True,fc=C_WHITE,size=9,align='center',wrap=True); c.border=_border()
    for col in [COL_CM_BUD,COL_CM_ACT]:
        c=ws.cell(8,col,MONTH_DATES[cm_idx]); c.number_format='MMM-YY'
        _s(c,bg=C_DARK_GREY,bold=True,fc=C_WHITE,size=9,align='center')
    for i in range(12):
        for base,bg in [(COL_BUD_START,C_DARK_GREEN),(COL_ACT_START,C_DARK_GREY),
                         (COL_VAR_START,C_DARK_GREY),(COL_INP_START,C_ORANGE)]:
            c=ws.cell(8,base+i,MONTH_DATES[i]); c.number_format='MMM-YY'
            _s(c,bg=bg,bold=True,fc=C_WHITE,size=8,align='center'); c.border=_border()

    # Group entries by Finance-facing category key.
    # For the "Performance Marketing" umbrella channel, split by finance_cat
    # (Meta, Bing, Douyin, LDD, etc.). For everything else, group by channel_name
    # (Affiliate - CPA & FF, Campaign/Promotions, Events, etc.) — matches the
    # original Finance template row structure.
    cats=OrderedDict()
    for e in entries:
        cat = _group_key(e)
        cats.setdefault(cat,[]).append(e)

    GRAND_ROW=9; current_row=10; cat_total_rows=[]

    for cat,cat_entries in cats.items():
        s_bg,d_bg=C_MID_GREEN,C_LIGHT_GREEN
        total_row=current_row; data_start=current_row+1; data_end=current_row+len(cat_entries)
        cat_total_rows.append(total_row)

        for col in range(1,TOTAL_COLS+1): _s(ws.cell(total_row,col),bg=s_bg,fc=C_WHITE)
        ws.cell(total_row,COL_CATEGORY,cat); _s(ws.cell(total_row,COL_CATEGORY),bg=s_bg,bold=True,fc=C_WHITE)
        ws.cell(total_row,COL_COUNTRY,"Total"); _s(ws.cell(total_row,COL_COUNTRY),bg=s_bg,bold=True,fc=C_WHITE)

        ws.cell(total_row,COL_FY26_ACT).value=f"=SUM({gl(COL_BUD_START)}{total_row}:{gl(COL_BUD_START+11)}{total_row})"
        ws.cell(total_row,COL_FY26_BUD).value=f"=SUM(H{data_start}:H{data_end})"
        ws.cell(total_row,COL_VAR_CHECK).value=f"=G{total_row}-H{total_row}"
        for fc in [COL_FY26_ACT,COL_FY26_BUD,COL_VAR_CHECK]:
            _s(ws.cell(total_row,fc),bg=s_bg,bold=True,fc=C_WHITE,nf=NUM_FMT)

        for i in range(12):
            for base in [COL_BUD_START,COL_ACT_START,COL_INP_START]:
                cl=gl(base+i); ws.cell(total_row,base+i).value=f"=SUM({cl}{data_start}:{cl}{data_end})"
                _s(ws.cell(total_row,base+i),bg=s_bg,bold=True,fc=C_WHITE,nf=NUM_FMT)
            ws.cell(total_row,COL_VAR_START+i).value=f"={gl(COL_BUD_START+i)}{total_row}-{gl(COL_ACT_START+i)}{total_row}"
            _s(ws.cell(total_row,COL_VAR_START+i),bg=s_bg,bold=True,fc=C_WHITE,nf=NUM_FMT)

        ws.cell(total_row,COL_YTD_BUD).value=f"=SUM({gl(COL_BUD_START)}{total_row}:{gl(COL_BUD_START+cm_idx)}{total_row})"
        ws.cell(total_row,COL_YTD_ACT).value=f"=SUM({gl(COL_ACT_START)}{total_row}:{gl(COL_ACT_START+cm_idx)}{total_row})"
        ws.cell(total_row,COL_YTD_VAR).value=f"={gl(COL_YTD_BUD)}{total_row}-{gl(COL_YTD_ACT)}{total_row}"
        ws.cell(total_row,COL_CM_BUD).value=f"={gl(COL_BUD_START+cm_idx)}{total_row}"
        ws.cell(total_row,COL_CM_ACT).value=f"={gl(COL_ACT_START+cm_idx)}{total_row}"
        ws.cell(total_row,COL_CM_VAR).value=f"={gl(COL_CM_BUD)}{total_row}-{gl(COL_CM_ACT)}{total_row}"
        for fc in [COL_YTD_BUD,COL_YTD_ACT,COL_YTD_VAR,COL_CM_BUD,COL_CM_ACT,COL_CM_VAR]:
            _s(ws.cell(total_row,fc),bg=s_bg,bold=True,fc=C_WHITE,nf=NUM_FMT)

        current_row+=1

        for j,entry in enumerate(cat_entries):
            row_bg=d_bg if j%2==0 else C_WHITE; r=current_row
            for col in range(1,TOTAL_COLS+1): _s(ws.cell(r,col),bg=row_bg)

            ws.cell(r,COL_ACCOUNT,entry.get("bu","")); ws.cell(r,COL_OWNER,entry.get("entered_by",""))
            ws.cell(r,COL_CATEGORY,cat); ws.cell(r,COL_COUNTRY,entry.get("country",""))
            ws.cell(r,COL_VENDOR,entry.get("vendor","")); ws.cell(r,COL_NOTE,entry.get("description","") or entry.get("activity_name",""))
            for col in [COL_ACCOUNT,COL_OWNER,COL_CATEGORY,COL_COUNTRY,COL_VENDOR,COL_NOTE]:
                _s(ws.cell(r,col),bg=row_bg,size=9)

            ws.cell(r,COL_FY26_ACT).value=f"=SUM({gl(COL_BUD_START)}{r}:{gl(COL_BUD_START+11)}{r})"
            _s(ws.cell(r,COL_FY26_ACT),bg=row_bg,bold=True,fc=C_MID_GREEN,nf=NUM_FMT)
            ws.cell(r,COL_FY26_BUD,float(entry.get("planned") or 0))
            _s(ws.cell(r,COL_FY26_BUD),bg=row_bg,fc=C_BLUE_INPUT,nf=NUM_FMT)
            ws.cell(r,COL_VAR_CHECK).value=f"=G{r}-H{r}"
            _s(ws.cell(r,COL_VAR_CHECK),bg=row_bg,nf=NUM_FMT)

            month_val=entry.get("month","")
            planned=float(entry.get("planned") or 0)
            actual=float(entry.get("actual") or 0)

            for i in range(12):
                is_entry_month=(MONTH_KEYS[i]==month_val)
                ws.cell(r,COL_BUD_START+i,planned if is_entry_month else 0)
                _s(ws.cell(r,COL_BUD_START+i),bg=row_bg,fc=C_BLUE_INPUT,nf=NUM_FMT)
                ws.cell(r,COL_BUD_START+i).border=_border("DDDDDD")
                ws.cell(r,COL_ACT_START+i,actual if is_entry_month else 0)
                _s(ws.cell(r,COL_ACT_START+i),bg=row_bg,fc=C_BLUE_INPUT,nf=NUM_FMT)
                ws.cell(r,COL_ACT_START+i).border=_border("DDDDDD")
                ws.cell(r,COL_VAR_START+i).value=f"={gl(COL_BUD_START+i)}{r}-{gl(COL_ACT_START+i)}{r}"
                _s(ws.cell(r,COL_VAR_START+i),bg=row_bg,nf=NUM_FMT)
                ws.cell(r,COL_INP_START+i,actual if is_entry_month else 0)
                _s(ws.cell(r,COL_INP_START+i),bg=row_bg,fc=C_ORANGE,nf=NUM_FMT)
                ws.cell(r,COL_INP_START+i).border=_border("DDDDDD")

            ws.cell(r,COL_YTD_BUD).value=f"=SUM({gl(COL_BUD_START)}{r}:{gl(COL_BUD_START+cm_idx)}{r})"
            ws.cell(r,COL_YTD_ACT).value=f"=SUM({gl(COL_ACT_START)}{r}:{gl(COL_ACT_START+cm_idx)}{r})"
            ws.cell(r,COL_YTD_VAR).value=f"={gl(COL_YTD_BUD)}{r}-{gl(COL_YTD_ACT)}{r}"
            ws.cell(r,COL_CM_BUD).value=f"={gl(COL_BUD_START+cm_idx)}{r}"
            ws.cell(r,COL_CM_ACT).value=f"={gl(COL_ACT_START+cm_idx)}{r}"
            ws.cell(r,COL_CM_VAR).value=f"={gl(COL_CM_BUD)}{r}-{gl(COL_CM_ACT)}{r}"
            for fc in [COL_YTD_BUD,COL_YTD_ACT,COL_YTD_VAR,COL_CM_BUD,COL_CM_ACT,COL_CM_VAR]:
                _s(ws.cell(r,fc),bg=row_bg,nf=NUM_FMT)

            ws.row_dimensions[r].height=16; current_row+=1

    for col in range(1,TOTAL_COLS+1): _s(ws.cell(GRAND_ROW,col),bg=C_DARK_GREEN)
    ws.cell(GRAND_ROW,COL_ACCOUNT,"AP- Main account"); _s(ws.cell(GRAND_ROW,COL_ACCOUNT),bg=C_DARK_GREEN,bold=True,fc=C_WHITE,size=9)
    ws.cell(GRAND_ROW,COL_OWNER,"Budget Owner"); _s(ws.cell(GRAND_ROW,COL_OWNER),bg=C_DARK_GREEN,bold=True,fc=C_WHITE,size=9)
    ws.cell(GRAND_ROW,COL_CATEGORY,"Total Marketing"); _s(ws.cell(GRAND_ROW,COL_CATEGORY),bg=C_DARK_GREEN,bold=True,fc=C_WHITE)
    ws.cell(GRAND_ROW,COL_COUNTRY,"Total"); _s(ws.cell(GRAND_ROW,COL_COUNTRY),bg=C_DARK_GREEN,bold=True,fc=C_WHITE)

    if cat_total_rows:
        for fc in [COL_FY26_ACT,COL_FY26_BUD,COL_VAR_CHECK,COL_YTD_BUD,COL_YTD_ACT,COL_YTD_VAR,COL_CM_BUD,COL_CM_ACT,COL_CM_VAR]:
            refs="+".join([f"{gl(fc)}{r}" for r in cat_total_rows])
            ws.cell(GRAND_ROW,fc).value=f"={refs}"
            _s(ws.cell(GRAND_ROW,fc),bg=C_DARK_GREEN,bold=True,fc=C_WHITE,nf=NUM_FMT)
        for i in range(12):
            for base in [COL_BUD_START,COL_ACT_START,COL_VAR_START,COL_INP_START]:
                col=base+i; refs="+".join([f"{gl(col)}{r}" for r in cat_total_rows])
                ws.cell(GRAND_ROW,col).value=f"={refs}"
                _s(ws.cell(GRAND_ROW,col),bg=C_DARK_GREEN,bold=True,fc=C_WHITE,nf=NUM_FMT)

    ws.row_dimensions[GRAND_ROW].height=22
    ws.freeze_panes='G10'

    nr=current_row+2
    ws.merge_cells(f'A{nr}:{gl(TOTAL_COLS)}{nr}')
    _s(ws.cell(nr,1,"Legend:  Blue = Budget input  |  Orange (BF-BQ) = Marketing team actual input  |  Black = Formulas  |  Var = Budget minus Actual"),italic=True,fc="666666",size=8)

    wb.save(output_path)
    return output_path
