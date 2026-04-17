# -*- coding: utf-8 -*-
"""
APAC Marketing Budget Tracker — Flask app entrypoint.

DATA ACCESS (for Postgres port): all storage calls go through sheets_helper
and the blueprint modules (api_pm, api_uploads). Swap sheets_helper.py with a
Postgres-backed version implementing the same function signatures.
See PORTING.md.

Run:  python app.py
"""
import os, json, uuid, base64, io, csv
from datetime import datetime
from flask import Flask, render_template, request, jsonify, session, redirect, send_file
from werkzeug.security import generate_password_hash, check_password_hash
from collections import defaultdict

from config import *
from sheets_helper import (get_sheet, safe_get_records, get_records_cached,
                           invalidate_cache, rows_for_cached, rows_for, ensure_entry_headers)
from auth import (seed_users, seed_categories, seed_mapping, get_user, get_all_users,
                  require_login, require_admin, check_country_access)
import api_pm
import api_uploads

app = Flask(__name__)
app.secret_key = SECRET_KEY
INVOICE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "invoices")
os.makedirs(INVOICE_DIR, exist_ok=True)
app.register_blueprint(api_pm.bp)
app.register_blueprint(api_uploads.bp)

def save_invoice_to_disk(data_url, entry_id, filename):
    try:
        header, b64 = data_url.split(",", 1); data = base64.b64decode(b64)
        safe_name = "".join(c for c in filename if c.isalnum() or c in "._- ")[:100]
        stored_name = f"{entry_id}_{uuid.uuid4().hex[:6]}_{safe_name}"
        with open(os.path.join(INVOICE_DIR, stored_name), "wb") as f: f.write(data)
        return stored_name
    except Exception: return None

def get_invoice_path(stored_name):
    path = os.path.join(INVOICE_DIR, stored_name)
    return path if os.path.exists(path) else None

@app.route("/")
def index():
    if not session.get("username"):
        users = get_all_users()
        return render_template("login.html", users=[{"username":u.get("username",""),"display_name":u.get("display_name",u.get("username",""))} for u in users], markets=MARKETS)
    role = session.get("role","country"); um = session.get("markets","")
    vm = [m.strip() for m in um.split(",")] if role=="country" and um!="ALL" else MARKETS
    return render_template("app.html", user=session.get("display_name",session["username"]), username=session["username"], is_admin=role=="admin", is_editor=role=="editor", role=role, markets=vm, quarters=QUARTERS)

@app.route("/login", methods=["POST"])
def login():
    username = request.form.get("username","").strip(); password = request.form.get("password","").strip()
    err_ctx = lambda msg: render_template("login.html", users=[{"username":u.get("username",""),"display_name":u.get("display_name","")} for u in get_all_users()], markets=MARKETS, error=msg)
    if not username or not password: return err_ctx("Enter username and password")
    user = get_user(username)
    if not user or not check_password_hash(str(user.get("password_hash","")), password): return err_ctx("Invalid password")
    session["username"]=user["username"]; session["display_name"]=user.get("display_name",user["username"])
    session["role"]=user.get("role","country"); session["markets"]=user.get("markets","ALL")
    session["user"]="APAC" if user.get("role")=="admin" else user.get("markets","").split(",")[0]
    return redirect("/")

@app.route("/logout")
def logout(): session.clear(); return redirect("/")

# -- CATEGORIES -----------------------------------------------------------
@app.route("/api/categories")
@require_login
def api_get_categories():
    try:
        rows = get_records_cached(TAB_CATEGORIES)
        if not rows: seed_categories(); invalidate_cache(TAB_CATEGORIES); rows = get_records_cached(TAB_CATEGORIES)
        result = {"bu":[],"finance":[],"marketing":[]}
        for r in rows:
            t = r.get("type","")
            if t in result: result[t].append({"id":r["id"],"value":r["value"],"sort_order":int(r.get("sort_order") or 0)})
        for t in result: result[t].sort(key=lambda x: x["sort_order"])
        return jsonify(result)
    except: return jsonify({"bu":DEFAULT_BU_LIST,"finance":DEFAULT_FIN_CATS,"marketing":DEFAULT_MKT_CATS})

@app.route("/api/categories", methods=["POST"])
@require_login
def api_add_category():
    # Admin OR editor can add categories (editors often need new finance/marketing
    # categories while entering line items). Country-role users cannot.
    role = session.get("role", "")
    if role not in ("admin", "editor"):
        return jsonify({"error":"Admin or editor required"}), 403
    d=request.get_json(); ct=d.get("type",""); v=d.get("value","").strip()
    if not ct or not v or ct not in ("bu","finance","marketing"): return jsonify({"error":"Invalid"}),400
    rows=get_records_cached(TAB_CATEGORIES)
    if any(r.get("type")==ct and str(r.get("value","")).lower()==v.lower() for r in rows): return jsonify({"error":"Exists"}),400
    cid=f"cat_{uuid.uuid4().hex[:8]}"; s=len([r for r in rows if r.get("type")==ct])
    get_sheet(TAB_CATEGORIES).append_row([cid,ct,v,s,datetime.utcnow().isoformat()]); invalidate_cache(TAB_CATEGORIES)
    return jsonify({"id":cid,"type":ct,"value":v,"sort_order":s})

@app.route("/api/categories/<cat_id>", methods=["DELETE"])
@require_login
@require_admin
def api_delete_category(cat_id):
    ws=get_sheet(TAB_CATEGORIES); rows=safe_get_records(ws,TAB_CATEGORIES)
    idx=next((i for i,r in enumerate(rows) if str(r.get("id",""))==cat_id),None)
    if idx is None: return jsonify({"error":"Not found"}),404
    ws.delete_rows(idx+2); invalidate_cache(TAB_CATEGORIES); return jsonify({"ok":True})

# -- BUDGET ---------------------------------------------------------------
@app.route("/api/budget/<country>/<quarter>")
@require_login
def api_get_budget(country, quarter):
    if not check_country_access(country): return jsonify({"error":"Forbidden"}),403
    try:
        brows=rows_for_cached(TAB_BUDGETS,country=country,quarter=quarter); total=float(brows[0]["total_budget"]) if brows else 0
        channels=sorted([{"id":r["id"],"name":r["name"],"budget":float(r["budget"] or 0),"sort_order":int(r.get("sort_order") or 0)} for r in rows_for_cached(TAB_CHANNELS,country=country,quarter=quarter)],key=lambda x:x["sort_order"])
        aa=rows_for_cached(TAB_ACTIVITIES,country=country,quarter=quarter)
        for ch in channels: ch["activities"]=sorted([{"id":a["id"],"name":a["name"],"sort_order":int(a.get("sort_order") or 0)} for a in aa if str(a["channel_id"])==str(ch["id"])],key=lambda x:x["sort_order"])
        try:
            mr=get_records_cached(TAB_MAPPING)
            if not mr: seed_mapping(); invalidate_cache(TAB_MAPPING); mr=get_records_cached(TAB_MAPPING)
            mapping=[{"channel_keyword":r["channel_keyword"],"bu":r["bu"],"finance_cat":r["finance_cat"],"marketing_cat":r["marketing_cat"]} for r in mr if r.get("channel_keyword")]
        except: mapping=[{"channel_keyword":kw,"bu":bu,"finance_cat":fc,"marketing_cat":mc} for kw,bu,fc,mc in DEFAULT_MAPPING]
        return jsonify({"total":total,"channels":channels,"mapping":mapping})
    except Exception as e: return jsonify({"error":str(e)}),500

@app.route("/api/budget/<country>/<quarter>", methods=["POST"])
@require_login
@require_admin
def api_save_budget(country, quarter):
    total=float(request.get_json().get("total",0)); ws=get_sheet(TAB_BUDGETS); rows=safe_get_records(ws,TAB_BUDGETS); now=datetime.utcnow().isoformat()
    idx=next((i for i,r in enumerate(rows) if r["country"]==country and r["quarter"]==quarter),None)
    if idx is not None: ws.update(f"A{idx+2}:E{idx+2}",[[rows[idx]["id"],country,quarter,total,now]])
    else: ws.append_row([str(uuid.uuid4())[:8],country,quarter,total,now])
    invalidate_cache(TAB_BUDGETS); return jsonify({"ok":True})

# -- CHANNELS -------------------------------------------------------------
@app.route("/api/channels", methods=["POST"])
@require_login
@require_admin
def api_add_channel():
    d=request.get_json(); ex=rows_for(TAB_CHANNELS,country=d["country"],quarter=d["quarter"]); cid="ch_"+str(uuid.uuid4())[:8]
    # Guard against exact-name duplicates (case-insensitive) — returns existing id instead
    existing = next((c for c in ex if str(c.get("name","")).strip().lower() == str(d["name"]).strip().lower()), None)
    if existing:
        return jsonify({"id":str(existing["id"]),"name":str(existing["name"]),
                        "budget":float(existing.get("budget") or 0),
                        "sort_order":int(existing.get("sort_order") or 0),
                        "duplicate":True})
    get_sheet(TAB_CHANNELS).append_row([cid,d["country"],d["quarter"],d["name"],float(d.get("budget",0)),len(ex),datetime.utcnow().isoformat()])
    invalidate_cache(TAB_CHANNELS); return jsonify({"id":cid,"name":d["name"],"budget":float(d.get("budget",0)),"sort_order":len(ex)})

@app.route("/api/channels/<ch_id>", methods=["PUT"])
@require_login
@require_admin
def api_update_channel(ch_id):
    d=request.get_json(); ws=get_sheet(TAB_CHANNELS); rows=safe_get_records(ws,TAB_CHANNELS)
    idx=next((i for i,r in enumerate(rows) if r["id"]==ch_id),None)
    if idx is None: return jsonify({"error":"Not found"}),404
    r=rows[idx]; ws.update(f"A{idx+2}:G{idx+2}",[[ch_id,r["country"],r["quarter"],d.get("name",r["name"]),float(d.get("budget",r["budget"])),r.get("sort_order",0),r.get("created_at","")]])
    invalidate_cache(TAB_CHANNELS); return jsonify({"ok":True})

@app.route("/api/channels/<ch_id>", methods=["DELETE"])
@require_login
@require_admin
def api_delete_channel(ch_id):
    ws=get_sheet(TAB_CHANNELS); rows=safe_get_records(ws,TAB_CHANNELS)
    idx=next((i for i,r in enumerate(rows) if r["id"]==ch_id),None)
    if idx is None: return jsonify({"error":"Not found"}),404
    ws.delete_rows(idx+2); invalidate_cache(TAB_CHANNELS); return jsonify({"ok":True})

# -- ACTIVITIES (all roles) -----------------------------------------------
@app.route("/api/activities", methods=["POST"])
@require_login
def api_add_activity():
    d=request.get_json(); ex=rows_for(TAB_ACTIVITIES,channel_id=d["channel_id"]); aid="act_"+str(uuid.uuid4())[:8]
    # Guard against duplicate (case-insensitive)
    existing = next((a for a in ex if str(a.get("name","")).strip().lower() == str(d["name"]).strip().lower()), None)
    if existing:
        return jsonify({"id":str(existing["id"]),"name":str(existing["name"]),
                        "sort_order":int(existing.get("sort_order") or 0),
                        "duplicate":True})
    get_sheet(TAB_ACTIVITIES).append_row([aid,d["channel_id"],d["country"],d["quarter"],d["name"],len(ex),datetime.utcnow().isoformat()])
    invalidate_cache(TAB_ACTIVITIES); return jsonify({"id":aid,"name":d["name"],"sort_order":len(ex)})

@app.route("/api/activities/<act_id>", methods=["PUT"])
@require_login
def api_update_activity(act_id):
    d=request.get_json(); ws=get_sheet(TAB_ACTIVITIES); rows=safe_get_records(ws,TAB_ACTIVITIES)
    idx=next((i for i,r in enumerate(rows) if r["id"]==act_id),None)
    if idx is None: return jsonify({"error":"Not found"}),404
    r=rows[idx]; ws.update(f"A{idx+2}:G{idx+2}",[[act_id,r["channel_id"],r["country"],r["quarter"],d.get("name",r["name"]),r.get("sort_order",0),r.get("created_at","")]])
    invalidate_cache(TAB_ACTIVITIES); return jsonify({"ok":True})

@app.route("/api/activities/<act_id>", methods=["DELETE"])
@require_login
def api_delete_activity(act_id):
    ws=get_sheet(TAB_ACTIVITIES); rows=safe_get_records(ws,TAB_ACTIVITIES)
    idx=next((i for i,r in enumerate(rows) if r["id"]==act_id),None)
    if idx is None: return jsonify({"error":"Not found"}),404
    ws.delete_rows(idx+2); invalidate_cache(TAB_ACTIVITIES); return jsonify({"ok":True})

# -- ENTRIES --------------------------------------------------------------
@app.route("/api/entries/<country>/<quarter>")
@require_login
def api_get_entries(country, quarter):
    if not check_country_access(country): return jsonify({"error":"Forbidden"}),403
    try:
        return jsonify([{"id":r["id"],"country":r["country"],"quarter":r["quarter"],"month":r["month"],"channel_id":r["channel_id"],"channel_name":r["channel_name"],"bu":r["bu"],"finance_cat":r["finance_cat"],"marketing_cat":r["marketing_cat"],"activity_id":r.get("activity_id",""),"activity_name":r.get("activity_name",""),"description":r["description"],"planned":float(r["planned"] or 0),"confirmed":float(r["confirmed"] or 0),"actual":float(r["actual"] or 0),"jira":r["jira"],"vendor":r["vendor"],"notes":r["notes"],"approved":str(r["approved"]).lower()=="true","invoice_names":json.loads(r["invoice_names"]) if r.get("invoice_names") else [],"entered_by":r["entered_by"],"updated_at":r["updated_at"]} for r in rows_for_cached(TAB_ENTRIES,country=country,quarter=quarter)])
    except Exception as e: return jsonify({"error":str(e)}),500

@app.route("/api/entries", methods=["POST"])
@require_login
def api_add_entry():
    d=request.get_json(); co=d.get("country","")
    if not check_country_access(co): return jsonify({"error":"Forbidden"}),403
    try:
        eid="e_"+str(uuid.uuid4())[:10]; now=datetime.utcnow().isoformat()
        ins=d.get("invoice_names",[]); ids=d.get("invoice_data",[])
        sf=[save_invoice_to_disk(ids[i],eid,n) or "" for i,n in enumerate(ins) if i<len(ids) and ids[i]]
        get_sheet(TAB_ENTRIES).append_row([eid,co,d.get("quarter",""),d.get("month",""),d.get("channel_id",""),d.get("channel_name",""),d.get("activity_id",""),d.get("activity_name",""),d.get("bu",""),d.get("finance_cat",""),d.get("marketing_cat",""),d.get("description",""),float(d.get("planned") or 0),float(d.get("confirmed") or 0),float(d.get("actual") or 0),d.get("jira",""),d.get("vendor",""),d.get("notes",""),str(d.get("approved",False)),json.dumps(ins),json.dumps(sf),session.get("username",""),now,now])
        invalidate_cache(TAB_ENTRIES); return jsonify({"id":eid,"ok":True})
    except Exception as e: return jsonify({"error":str(e)}),500

@app.route("/api/entries/<entry_id>", methods=["PUT"])
@require_login
def api_update_entry(entry_id):
    d=request.get_json()
    try:
        ws=get_sheet(TAB_ENTRIES); rows=safe_get_records(ws,TAB_ENTRIES)
        idx=next((i for i,r in enumerate(rows) if str(r.get("id",""))==str(entry_id)),None)
        if idx is None: return jsonify({"error":"Not found"}),404
        r=rows[idx]
        if not check_country_access(str(r.get("country",""))): return jsonify({"error":"Forbidden"}),403
        en=json.loads(str(r.get("invoice_names") or "[]")); ef=json.loads(str(r.get("invoice_data") or "[]"))
        inn=d.get("invoice_names",None); ind=d.get("invoice_data",None)
        if inn is not None:
            nu=ind if ind else []; fn=[]; ff=[]
            for nm in inn:
                if nm in en: oi=en.index(nm); fn.append(nm); ff.append(ef[oi] if oi<len(ef) else ""); en[oi]=None
                elif nu: fn.append(nm); ff.append(save_invoice_to_disk(nu.pop(0),entry_id,nm) or "")
                else: fn.append(nm); ff.append("")
            inv_n,inv_d=json.dumps(fn),json.dumps(ff)
        else: inv_n,inv_d=json.dumps(en),json.dumps(ef)
        now=datetime.utcnow().isoformat()
        ws.update(f"A{idx+2}:X{idx+2}",[[entry_id,r.get("country",""),d.get("quarter",r.get("quarter","")),d.get("month",r.get("month","")),d.get("channel_id",r.get("channel_id","")),d.get("channel_name",r.get("channel_name","")),d.get("activity_id",r.get("activity_id","")),d.get("activity_name",r.get("activity_name","")),d.get("bu",r.get("bu","")),d.get("finance_cat",r.get("finance_cat","")),d.get("marketing_cat",r.get("marketing_cat","")),d.get("description",r.get("description","")),float(d.get("planned",r.get("planned",0)) or 0),float(d.get("confirmed",r.get("confirmed",0)) or 0),float(d.get("actual",r.get("actual",0)) or 0),d.get("jira",r.get("jira","")),d.get("vendor",r.get("vendor","")),d.get("notes",r.get("notes","")),str(d.get("approved",str(r.get("approved","")).lower()=="true")),inv_n,inv_d,r.get("entered_by",""),r.get("created_at",""),now]])
        invalidate_cache(TAB_ENTRIES); return jsonify({"ok":True})
    except Exception as e: return jsonify({"error":str(e)}),500

@app.route("/api/entries/<entry_id>", methods=["DELETE"])
@require_login
def api_delete_entry(entry_id):
    ws=get_sheet(TAB_ENTRIES); rows=safe_get_records(ws,TAB_ENTRIES)
    idx=next((i for i,r in enumerate(rows) if str(r.get("id",""))==str(entry_id)),None)
    if idx is None: return jsonify({"error":"Not found"}),404
    if session["user"]!=ADMIN_MARKET and str(rows[idx].get("country",""))!=session["user"]: return jsonify({"error":"Forbidden"}),403
    ws.delete_rows(idx+2); invalidate_cache(TAB_ENTRIES); return jsonify({"ok":True})

# -- VENDORS --------------------------------------------------------------
@app.route("/api/vendors")
@require_login
def api_get_vendors():
    try:
        vs=get_records_cached(TAB_VENDORS); u=session.get("user","")
        if u!=ADMIN_MARKET: vs=[v for v in vs if v.get("country") in ("GLOBAL",u)]
        seen=set(); ul=[]
        for v in [{"id":v["id"],"name":v["name"],"country":v.get("country","GLOBAL")} for v in vs]:
            if v["name"].lower() not in seen: seen.add(v["name"].lower()); ul.append(v)
        return jsonify(sorted(ul,key=lambda x:x["name"].lower()))
    except: return jsonify([])

@app.route("/api/vendors", methods=["POST"])
@require_login
def api_add_vendor():
    d=request.get_json(); nm=d.get("name","").strip()
    if not nm: return jsonify({"error":"Required"}),400
    u=session.get("user",""); vc=d.get("country","GLOBAL" if u==ADMIN_MARKET else u)
    vid="v_"+str(uuid.uuid4())[:8]
    get_sheet(TAB_VENDORS).append_row([vid,nm,vc,u,datetime.utcnow().isoformat()]); invalidate_cache(TAB_VENDORS)
    return jsonify({"id":vid,"name":nm,"country":vc})

# -- USERS ----------------------------------------------------------------
@app.route("/api/users")
@require_login
@require_admin
def api_get_users():
    return jsonify([{"username":u.get("username",""),"display_name":u.get("display_name",""),"role":u.get("role","country"),"markets":u.get("markets","")} for u in get_all_users()])

@app.route("/api/users", methods=["POST"])
@require_login
@require_admin
def api_add_user():
    d=request.get_json(); un=d.get("username","").strip().lower(); pw=d.get("password","").strip()
    if not un or not pw: return jsonify({"error":"Required"}),400
    if get_user(un): return jsonify({"error":"Exists"}),400
    get_sheet(TAB_USERS).append_row([un,generate_password_hash(pw),d.get("display_name","").strip() or un,d.get("role","country"),d.get("markets","ALL"),datetime.utcnow().isoformat()])
    invalidate_cache(TAB_USERS); return jsonify({"ok":True,"username":un})

@app.route("/api/users/<username>", methods=["DELETE"])
@require_login
@require_admin
def api_delete_user(username):
    if username.lower()==session.get("username","").lower(): return jsonify({"error":"Cannot delete self"}),400
    ws=get_sheet(TAB_USERS); rows=safe_get_records(ws,TAB_USERS)
    idx=next((i for i,r in enumerate(rows) if str(r.get("username","")).lower()==username.lower()),None)
    if idx is None: return jsonify({"error":"Not found"}),404
    ws.delete_rows(idx+2); invalidate_cache(TAB_USERS); return jsonify({"ok":True})

# -- MAPPING --------------------------------------------------------------
@app.route("/api/mapping")
@require_login
def api_get_mapping():
    try:
        mr=get_records_cached(TAB_MAPPING)
        if not mr: seed_mapping(); invalidate_cache(TAB_MAPPING); mr=get_records_cached(TAB_MAPPING)
        return jsonify([{"channel_keyword":r["channel_keyword"],"bu":r["bu"],"finance_cat":r["finance_cat"],"marketing_cat":r["marketing_cat"]} for r in mr if r.get("channel_keyword")])
    except: return jsonify([{"channel_keyword":kw,"bu":bu,"finance_cat":fc,"marketing_cat":mc} for kw,bu,fc,mc in DEFAULT_MAPPING])

# -- RECONCILIATION -------------------------------------------------------
@app.route("/api/reconciliation/<quarter>")
@require_login
def api_reconciliation(quarter):
    try:
        role=session.get("role",""); um=session.get("markets","")
        allowed=None if role in ("admin","editor") or um=="ALL" else set(m.strip() for m in um.split(",") if m.strip())
        ab=get_records_cached(TAB_BUDGETS); ach=get_records_cached(TAB_CHANNELS); aact=get_records_cached(TAB_ACTIVITIES)
        qb=[b for b in ab if str(b.get("quarter",""))==quarter]; qe=rows_for_cached(TAB_ENTRIES,quarter=quarter)
        mkts=sorted({str(b["country"]) for b in qb}|{str(e["country"]) for e in qe})
        if allowed: mkts=[m for m in mkts if m in allowed]
        result=[]
        for mkt in mkts:
            br=next((b for b in qb if str(b["country"])==mkt),None); pb=float(br["total_budget"]) if br else 0
            mc=[c for c in ach if str(c.get("country",""))==mkt and str(c.get("quarter",""))==quarter]
            me=rows_for_cached(TAB_ENTRIES,country=mkt,quarter=quarter); cd=[]; asgn=set()
            for ch in sorted(mc,key=lambda c:int(c.get("sort_order") or 0)):
                cid=str(ch["id"]); ce=[e for e in me if str(e.get("channel_id",""))==cid]
                ca=[a for a in aact if str(a.get("channel_id",""))==cid and str(a.get("country",""))==mkt and str(a.get("quarter",""))==quarter]
                ad=[]; aids=set()
                for act in sorted(ca,key=lambda a:int(a.get("sort_order") or 0)):
                    ae=[e for e in ce if str(e.get("activity_id",""))==str(act["id"])]; items=[]
                    for e in ae: aids.add(str(e["id"])); asgn.add(str(e["id"])); items.append({"id":e["id"],"month":e.get("month",""),"description":e.get("description",""),"vendor":e.get("vendor",""),"planned":float(e.get("planned") or 0),"confirmed":float(e.get("confirmed") or 0),"actual":float(e.get("actual") or 0),"jira":e.get("jira",""),"approved":str(e.get("approved","")).lower()=="true"})
                    ad.append({"id":act["id"],"name":act["name"],"planned":sum(i["planned"] for i in items),"confirmed":sum(i["confirmed"] for i in items),"actual":sum(i["actual"] for i in items),"entries":len(items),"items":items})
                ua=[e for e in ce if str(e["id"]) not in aids]
                for e in ua: asgn.add(str(e["id"]))
                ui=[{"id":e["id"],"month":e.get("month",""),"description":e.get("description",""),"vendor":e.get("vendor",""),"planned":float(e.get("planned") or 0),"confirmed":float(e.get("confirmed") or 0),"actual":float(e.get("actual") or 0),"jira":e.get("jira",""),"approved":str(e.get("approved","")).lower()=="true"} for e in ua]
                cd.append({"id":ch["id"],"name":ch["name"],"budget":float(ch.get("budget") or 0),"planned":sum(float(e.get("planned") or 0) for e in ce),"confirmed":sum(float(e.get("confirmed") or 0) for e in ce),"actual":sum(float(e.get("actual") or 0) for e in ce),"entries":len(ce),"activities":ad,"unassigned":ui})
            oe=[{"id":e["id"],"month":e.get("month",""),"description":e.get("description",""),"vendor":e.get("vendor",""),"planned":float(e.get("planned") or 0),"actual":float(e.get("actual") or 0)} for e in me if str(e["id"]) not in asgn]
            sp=sum(float(e.get("planned") or 0) for e in me); sa=sum(float(e.get("actual") or 0) for e in me)
            # Manual entries = non-PM-sync. PM-synced rows have id prefix "pm_".
            # Compliance math excludes PM rows from "manual" count since they're auto-generated.
            manual = [e for e in me if not str(e.get("id","")).startswith("pm_")]
            auto = [e for e in me if str(e.get("id","")).startswith("pm_")]
            result.append({
                "country":mkt,"plan_budget":pb,
                "sum_planned":sp,
                "sum_confirmed":sum(float(e.get("confirmed") or 0) for e in me),
                "sum_actual":sa,
                "entries":len(me),
                "entries_manual":len(manual),
                "entries_auto":len(auto),
                "channels_data":cd,"orphan_entries":oe,
                "var_plan_vs_actual":sa-sp,
                "flags":{
                    "no_actual":sum(1 for e in me if float(e.get("planned") or 0)>0 and float(e.get("actual") or 0)==0),
                    "no_jira":sum(1 for e in me if not e.get("jira")),
                    "no_invoice":sum(1 for e in me if float(e.get("actual") or 0)>0 and not str(e.get("invoice_names","[]")).strip("[]")),
                    # Manual-only counts — exclude PM sync rows because they don't need JIRA tickets
                    "no_jira_manual":sum(1 for e in manual if not e.get("jira")),
                    "no_invoice_manual":sum(1 for e in manual if float(e.get("actual") or 0)>0 and not str(e.get("invoice_names","[]")).strip("[]")),
                }
            })
        return jsonify(result)
    except Exception as e:
        import traceback; traceback.print_exc(); return jsonify({"error":str(e)}),500

# -- ANALYTICS ------------------------------------------------------------
@app.route("/api/analytics")
@require_login
def api_analytics():
    try:
        qf=request.args.get("quarter",""); cf=request.args.get("country","")
        ae=get_records_cached(TAB_ENTRIES); ab=get_records_cached(TAB_BUDGETS); ac=get_records_cached(TAB_CHANNELS)
        if qf: ae=[e for e in ae if str(e.get("quarter",""))==qf]; ab=[b for b in ab if str(b.get("quarter",""))==qf]; ac=[c for c in ac if str(c.get("quarter",""))==qf]
        if cf: ae=[e for e in ae if str(e.get("country",""))==cf]; ab=[b for b in ab if str(b.get("country",""))==cf]; ac=[c for c in ac if str(c.get("country",""))==cf]
        tb=sum(float(b.get("total_budget") or 0) for b in ab); tp=sum(float(e.get("planned") or 0) for e in ae); tc=sum(float(e.get("confirmed") or 0) for e in ae); ta=sum(float(e.get("actual") or 0) for e in ae)
        bc=defaultdict(lambda:{"budget":0,"planned":0,"confirmed":0,"actual":0,"entries":0})
        for b in ab: bc[str(b["country"])]["budget"]+=float(b.get("total_budget") or 0)
        for e in ae: c=str(e.get("country","")); bc[c]["planned"]+=float(e.get("planned") or 0); bc[c]["confirmed"]+=float(e.get("confirmed") or 0); bc[c]["actual"]+=float(e.get("actual") or 0); bc[c]["entries"]+=1
        bch=defaultdict(lambda:{"planned":0,"confirmed":0,"actual":0,"entries":0,"budget":0})
        for c in ac: bch[str(c.get("name",""))]["budget"]+=float(c.get("budget") or 0)
        for e in ae: ch=str(e.get("channel_name","")) or "Other"; bch[ch]["planned"]+=float(e.get("planned") or 0); bch[ch]["confirmed"]+=float(e.get("confirmed") or 0); bch[ch]["actual"]+=float(e.get("actual") or 0); bch[ch]["entries"]+=1
        bm=defaultdict(lambda:{"planned":0,"confirmed":0,"actual":0,"entries":0})
        for e in ae:
            mo=str(e.get("month",""))
            if mo: bm[mo]["planned"]+=float(e.get("planned") or 0); bm[mo]["confirmed"]+=float(e.get("confirmed") or 0); bm[mo]["actual"]+=float(e.get("actual") or 0); bm[mo]["entries"]+=1
        bmc=defaultdict(lambda:{"planned":0,"actual":0,"entries":0})
        for e in ae: mc=str(e.get("marketing_cat","")) or "Other"; bmc[mc]["planned"]+=float(e.get("planned") or 0); bmc[mc]["actual"]+=float(e.get("actual") or 0); bmc[mc]["entries"]+=1
        var=[{"id":e["id"],"country":e["country"],"channel":e.get("channel_name",""),"description":e.get("description","") or e.get("activity_name",""),"planned":float(e.get("planned") or 0),"actual":float(e.get("actual") or 0),"variance":float(e.get("actual") or 0)-float(e.get("planned") or 0),"variance_pct":((float(e.get("actual") or 0)-float(e.get("planned") or 0))/float(e["planned"])*100) if float(e.get("planned") or 0)>0 else 0} for e in ae if float(e.get("planned") or 0)>0 or float(e.get("actual") or 0)>0]
        var.sort(key=lambda x:abs(x["variance"]),reverse=True)
        wa=sum(1 for e in ae if float(e.get("actual") or 0)>0); wj=sum(1 for e in ae if e.get("jira")); ap=sum(1 for e in ae if str(e.get("approved","")).lower()=="true")
        MC=["HKG","CN","TW","TH","VN","SG","MY","MN","IN","APAC","ID","PH"]
        chm={}
        for e in ae: cid=str(e.get("channel_id","")); mc=str(e.get("marketing_cat","")); chm[cid]=mc if cid and mc else chm.get(cid,"")
        bbcm=defaultdict(lambda:defaultdict(float)); pbcm=defaultdict(lambda:defaultdict(float)); abcm=defaultdict(lambda:defaultdict(float)); amcs=set()
        for c in ac:
            cid=str(c.get("id","")); co=str(c.get("country","")); mc=chm.get(cid,"")
            if mc and co: bbcm[mc][co]+=float(c.get("budget") or 0)
        ctb=defaultdict(float)
        for b in ab: ctb[str(b.get("country",""))]+=float(b.get("total_budget") or 0)
        for e in ae: mc=str(e.get("marketing_cat","")) or "Other"; co=str(e.get("country","")); pbcm[mc][co]+=float(e.get("planned") or 0); abcm[mc][co]+=float(e.get("actual") or 0); amcs.add(mc)
        for mc in bbcm: amcs.add(mc)
        mx=[]
        for mc in sorted(amcs):
            row={"category":mc}; tb2=tp2=ta2=0
            for co in MC: b=bbcm[mc].get(co,0); p=pbcm[mc].get(co,0); a=abcm[mc].get(co,0); row[co+"_bud"]=b; row[co+"_pln"]=p; row[co+"_act"]=a; tb2+=b; tp2+=p; ta2+=a
            row["total_bud"]=tb2; row["total_pln"]=tp2; row["total_act"]=ta2
            if tb2>0 or tp2>0 or ta2>0: mx.append(row)
        tr={"category":"Total"}; gb=gp=ga=0
        for co in MC: cb=ctb.get(co,0); cp=sum(pbcm[mc].get(co,0) for mc in amcs); ca=sum(abcm[mc].get(co,0) for mc in amcs); tr[co+"_bud"]=cb; tr[co+"_pln"]=cp; tr[co+"_act"]=ca; gb+=cb; gp+=cp; ga+=ca
        tr["total_bud"]=gb; tr["total_pln"]=gp; tr["total_act"]=ga; mx.append(tr)
        return jsonify({"summary":{"total_budget":tb,"total_planned":tp,"total_confirmed":tc,"total_actual":ta,"total_entries":len(ae),"variance":ta-tp,"budget_utilization":(ta/tb*100) if tb>0 else 0},"by_country":[{"country":k,**v} for k,v in sorted(bc.items())],"by_channel":[{"channel":k,**v} for k,v in sorted(bch.items(),key=lambda x:-x[1]["actual"])],"by_month":[{"month":k,**v} for k,v in sorted(bm.items())],"by_marketing_cat":[{"category":k,**v} for k,v in sorted(bmc.items(),key=lambda x:-x[1]["actual"])],"top_variances":var[:20],"completion":{"total":len(ae),"with_actual":wa,"with_jira":wj,"approved":ap},"budget_matrix":mx,"matrix_countries":MC})
    except Exception as e:
        import traceback; traceback.print_exc(); return jsonify({"error":str(e)}),500

# -- EXPORT ---------------------------------------------------------------
@app.route("/api/export")
@require_login
def api_export():
    u=session["user"]; rows=safe_get_records(get_sheet(TAB_ENTRIES),TAB_ENTRIES)
    if u!=ADMIN_MARKET: rows=[r for r in rows if r["country"]==u]
    out=io.StringIO(); w=csv.writer(out)
    w.writerow(["Country","Quarter","Month","Channel","Activity","BU","Finance Category","Marketing Category","Description","Planned","Confirmed","Actual","JIRA","Vendor","Notes","Approved","Entered By","Updated At"])
    for r in rows: w.writerow([r["country"],r["quarter"],r["month"],r["channel_name"],r.get("activity_name",""),r["bu"],r["finance_cat"],r["marketing_cat"],r["description"],r["planned"],r["confirmed"],r["actual"],r["jira"],r["vendor"],r["notes"],"Yes" if str(r["approved"])=="True" else "No",r["entered_by"],r["updated_at"]])
    return send_file(io.BytesIO(out.getvalue().encode("utf-8-sig")),mimetype="text/csv",as_attachment=True,download_name=f"APAC_Budget_{'ALL' if u==ADMIN_MARKET else u}_{datetime.now().strftime('%Y-%m-%d')}.csv")

@app.route("/api/export/xlsx")
@require_login
def api_export_xlsx():
    from export_xlsx import build_finance_export; import tempfile
    u=session["user"]; ae=safe_get_records(get_sheet(TAB_ENTRIES),TAB_ENTRIES); ac=safe_get_records(get_sheet(TAB_CHANNELS),TAB_CHANNELS); ab=safe_get_records(get_sheet(TAB_BUDGETS),TAB_BUDGETS)
    if u!=ADMIN_MARKET: ae=[e for e in ae if str(e.get("country",""))==u]; ac=[c for c in ac if str(c.get("country",""))==u]; ab=[b for b in ab if str(b.get("country",""))==u]
    tmp=tempfile.NamedTemporaryFile(suffix=".xlsx",delete=False); tmp.close()
    try: build_finance_export(tmp.name,ae,ac,ab)
    except Exception as ex: os.unlink(tmp.name); import traceback; traceback.print_exc(); return jsonify({"error":f"Export failed: {ex}"}),500
    resp=send_file(tmp.name,mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",as_attachment=True,download_name=f"APAC_FY26_{'ALL' if u==ADMIN_MARKET else u}_{datetime.now().strftime('%Y-%m-%d')}.xlsx")
    @resp.call_on_close
    def cleanup():
        try: os.unlink(tmp.name)
        except: pass
    return resp

@app.route("/api/invoice/<entry_id>/<int:inv_idx>")
@require_login
def api_invoice(entry_id, inv_idx):
    rows=safe_get_records(get_sheet(TAB_ENTRIES),TAB_ENTRIES); r=next((row for row in rows if row["id"]==entry_id),None)
    if not r: return "Not found",404
    names=json.loads(r.get("invoice_names") or "[]"); datas=json.loads(r.get("invoice_data") or "[]")
    if inv_idx>=len(datas): return "Not found",404
    name=names[inv_idx] if inv_idx<len(names) else f"invoice_{inv_idx}"; stored=datas[inv_idx]
    if stored and not stored.startswith("data:"):
        path=get_invoice_path(stored)
        if path: import mimetypes; return send_file(path,mimetype=mimetypes.guess_type(name)[0] or "application/octet-stream",as_attachment=True,download_name=name)
    if stored and stored.startswith("data:"):
        try: h,b64=stored.split(",",1); return send_file(io.BytesIO(base64.b64decode(b64)),mimetype=h.split(";")[0].replace("data:",""),as_attachment=True,download_name=name)
        except: pass
    return "Not found",404

# -- LEGACY BULK IMPORT (kept — used by existing Config page) --------------
@app.route("/api/import/channels", methods=["POST"])
@require_login
@require_admin
def api_import_channels():
    if 'file' not in request.files: return jsonify({"error":"No file"}),400
    f=request.files['file']; parsed=[]
    try:
        if f.filename.lower().endswith(('.xlsx','.xls')):
            import openpyxl; wb=openpyxl.load_workbook(io.BytesIO(f.read()),data_only=True,read_only=True); ws=wb.active; ar=list(ws.iter_rows(values_only=True)); wb.close()
            st=1 if ar and any(k in ' '.join([str(v or '').lower() for v in ar[0]]) for k in ['country','quarter','channel']) else 0
            parsed=[[str(v).strip() if v else '' for v in r] for r in ar[st:] if any(str(v or '').strip() for v in r)]
        else:
            lines=[l.strip() for l in f.read().decode('utf-8-sig').splitlines() if l.strip()]
            st=1 if lines and any(k in lines[0].lower() for k in ['country','quarter','channel']) else 0
            import csv as _c; parsed=list(_c.reader(lines[st:]))
    except Exception as e: return jsonify({"error":str(e)}),400
    wc=get_sheet(TAB_CHANNELS); wb2=get_sheet(TAB_BUDGETS); ec=safe_get_records(wc,TAB_CHANNELS); eb=safe_get_records(wb2,TAB_BUDGETS); now=datetime.utcnow().isoformat(); saved=skipped_dups=0; sr=[]
    for row in parsed:
        if len(row)<3: continue
        co,q,nm=row[0].strip(),row[1].strip().upper(),row[2].strip(); bv=float(str(row[3]).replace(',','').replace('$','').strip() or 0) if len(row)>=4 else 0
        if not co or not q or not nm: continue
        if not q.startswith('Q'): q='Q'+q
        if not any(r['country']==co and r['quarter']==q for r in eb): wb2.append_row([str(uuid.uuid4())[:8],co,q,0,now]); eb.append({'country':co,'quarter':q,'total_budget':0})
        # Case-insensitive dedup check
        if any(r['country']==co and r['quarter']==q and str(r['name']).strip().lower()==nm.lower() for r in ec): skipped_dups+=1; continue
        cid="ch_"+str(uuid.uuid4())[:8]; so=len([r for r in ec if r['country']==co and r['quarter']==q])
        wc.append_row([cid,co,q,nm,bv,so,now]); ec.append({'id':cid,'country':co,'quarter':q,'name':nm,'budget':bv}); saved+=1; sr.append({"country":co,"quarter":q,"name":nm,"budget":bv})
    invalidate_cache(TAB_CHANNELS); invalidate_cache(TAB_BUDGETS)
    return jsonify({"ok":True,"saved":saved,"skipped":skipped_dups,"skipped_dups":skipped_dups,"rows":sr})

@app.route("/api/import/budgets", methods=["POST"])
@require_login
@require_admin
def api_import_budgets():
    if 'file' not in request.files: return jsonify({"error":"No file"}),400
    f=request.files['file']; parsed=[]
    try:
        if f.filename.lower().endswith(('.xlsx','.xls')):
            import openpyxl; wb=openpyxl.load_workbook(io.BytesIO(f.read()),data_only=True,read_only=True); ws=wb.active; ar=list(ws.iter_rows(values_only=True)); wb.close()
            st=1 if ar and any(k in ' '.join([str(v or '').lower() for v in ar[0]]) for k in ['country','quarter','budget']) else 0
            parsed=[[str(v).strip() if v else '' for v in r] for r in ar[st:] if any(str(v or '').strip() for v in r)]
        else:
            lines=[l.strip() for l in f.read().decode('utf-8-sig').splitlines() if l.strip()]
            st=1 if lines and any(k in lines[0].lower() for k in ['country','quarter','budget']) else 0
            import csv as _c; parsed=list(_c.reader(lines[st:]))
    except Exception as e: return jsonify({"error":str(e)}),400
    ws=get_sheet(TAB_BUDGETS); ex=safe_get_records(ws,TAB_BUDGETS); now=datetime.utcnow().isoformat(); saved=skipped=0; sr=[]
    for row in parsed:
        if len(row)<3: skipped+=1; continue
        co=row[0].strip(); q=row[1].strip().upper()
        if not q.startswith('Q'): q='Q'+q
        try: tot=float(str(row[2]).replace(',','').replace('$','').strip() or 0)
        except: skipped+=1; continue
        idx=next((i for i,r in enumerate(ex) if r['country']==co and r['quarter']==q),None)
        if idx is not None: ws.update(f"A{idx+2}:E{idx+2}",[[ex[idx]['id'],co,q,tot,now]])
        else: ws.append_row([str(uuid.uuid4())[:8],co,q,tot,now]); ex.append({'country':co,'quarter':q,'total_budget':tot})
        saved+=1; sr.append({"country":co,"quarter":q,"total":tot})
    invalidate_cache(TAB_BUDGETS); return jsonify({"ok":True,"saved":saved,"skipped":skipped,"rows":sr})

# Legacy template endpoints (kept for existing UI compatibility)
@app.route("/api/budget_template")
@require_login
def api_budget_template():
    out=io.StringIO(); w=csv.writer(out)
    w.writerow(["Country","Quarter","Total Budget (AUD)"])
    w.writerow(["TH","Q1","150000"]); w.writerow(["HKG","Q1","180000"])
    return send_file(io.BytesIO(out.getvalue().encode("utf-8-sig")),mimetype="text/csv",as_attachment=True,download_name="budget_template.csv")

@app.route("/api/channel_template")
@require_login
def api_channel_template():
    out=io.StringIO(); w=csv.writer(out)
    w.writerow(["Country","Quarter","Channel Name","Budget (AUD)"])
    w.writerow(["TH","Q1","Performance Marketing","80000"])
    w.writerow(["TH","Q1","Affiliate - CPA & FF","10000"])
    w.writerow(["TH","Q1","Campaign/Promotions","15000"])
    return send_file(io.BytesIO(out.getvalue().encode("utf-8-sig")),mimetype="text/csv",as_attachment=True,download_name="channel_template.csv")

if __name__ == "__main__":
    try: ensure_entry_headers()
    except Exception as e: print(f"[STARTUP] {e}")
    try: seed_users()
    except Exception as e: print(f"[STARTUP] {e}")
    try: seed_categories()
    except Exception as e: print(f"[STARTUP] {e}")
    app.run(debug=True, port=5000)
